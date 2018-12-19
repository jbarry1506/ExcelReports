from os import listdir
from os.path import (
    isfile,
    join,
)
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment,
)
import win32com.client as win32
import psutil
import os
import subprocess
import company_dictionary


def copy_cells(startrow, endrow, startcol, endcol, sheet):
    range_sel = []
    for i in range(startrow, endrow + 1, 1):
        rowsel = []
        for j in range(startcol, endcol + 1, 1):
            rowsel.append(ws.cell(row=i, column=j).value)
        range_sel.append(rowsel)
    return range_sel


def create_worksheet(workbook):
    new_sheet_names = workbook.get_sheet_names()
    nws = workbook.get_sheet_by_name(new_sheet_names[0])
    nws.title = "Ticket info"
    return nws


# put all files from a directory into an array
def file_array(f_dir):
    files = [i for i in listdir(path=f_dir) if isfile(join(f_dir, i))]
    for i in files:
        print(i)
    return files


def get_company_names(file):
    fn = open(file, 'r')
    comp_array = []
    for line in fn.readlines():
        comp_array.append(line)
    return comp_array


def find_text(s_row, e_row, s_col, e_col, searchsheet, searchtext):
    cell_info = copy_cells(s_row, e_row, s_col, e_col, searchsheet)
    row = 0
    for c in cell_info:
        row += 1
        for v in c:
            if v == searchtext:
                return row


def format_cell_width(sheet):
    sheet.column_dimensions['A'].width = 31
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['F'].width = 20


def format_heading(r, c):
    tws_cell = tws.cell(r, c)
    tws_cell.font = font_heading
    tws_cell.fill = fill_heading
    tws_cell.border = border
    tws_cell.alignment = alignment_center


def format_sub(r, c):
    tws_cell = tws.cell(r, c)
    tws_cell.font = font_sub
    tws_cell.border = border
    tws_cell.alignment = alignment_center


def format_sub_bold(r, c):
    tws_cell = tws.cell(r, c)
    tws_cell.font = font_sub_bold
    tws_cell.border = border
    tws_cell.alignment = alignment_center


def format_sub_gray(r, c):
    tws_cell = tws.cell(r, c)
    tws_cell.font = font_sub
    tws_cell.fill = fill_ticket
    tws_cell.border = border
    tws_cell.alignment = alignment_center


def format_title(r, c):
    tws_cell = tws.cell(r, c)
    tws_cell.font = font_title
    tws_cell.border = border
    tws_cell.alignment = alignment_center


# Open Outlook.exe. Path may vary according to system config
# Please check the path to .exe file and update below
def open_outlook():
    try:
        subprocess.call([r'  <  YOUR PATH  >  '])
        os.system(r'  <  YOUR PATH  >  ')
    except:
        print("Outlook didn't open successfully")


# Open the original workbook and sheet to get data from
def open_worksheet(file):
    wb = openpyxl.load_workbook(file)
    sheet_names = wb.get_sheet_names()
    my_sheet = wb.get_sheet_by_name(sheet_names[0])
    return my_sheet


def paste_cells(startrow, endrow, startcol, endcol, newsheet, data):
    rowcount = 0
    title_tuple = (
        'Closed',
        'In Progress-Assigned Engineer',
        'Monitoring',
        'Threshold Review',
        'Waiting on Customer',
        'Pending Change Order',
        'Closed - Send Survey',
        'Scheduled'
    )
    heading_tuple = (
        'Ticket Number',
        'Summary',
        'Issue Type',
        'Priority',
        'Contact',
        'Date Closed'
    )
    for i in range(startrow, endrow+1, 1):
        colcount = 0
        for j in range(startcol, endcol):
            newsheet.cell(row=i, column=j).value = data[rowcount][colcount]
            # print(newsheet.cell(i, j).value)

            if str(newsheet.cell(row=i, column=j).value) in heading_tuple:
                format_heading(i, j)
            else:
                format_sub(i, j)
                if j == 1:
                    if str(newsheet.cell(row=i, column=j).value) in title_tuple:
                        format_title(i, j)

            colcount += 1
        rowcount += 1


def save_sheet(workbook, file, company):
    # Create a new sheet to paste data
    save_name = str(company) + "_GEMS_Report_jimtest.xlsx"
    print("save_name " + save_name)
    save_filename = join(file, save_name)
    print("save_filename " + save_filename)

    workbook.save(save_filename)


def send_notification(file, company):
    # TODO - Write message body function
    def _message_body(cmpny):
        return {

        }

    def _cc_email(cmpny):
        return company_dictionary.email_lists.get(cmpny).get("CC")

    def _select_email(cmpny):
        return company_dictionary.email_lists.get(cmpny).get("To")

    # TODO Write subject string function
    def _subject_string(cmpny):
        return {

        }

    cc_list = [
        # < THIS IS A LIST OF ADDITIONAL EMAILS TO INCLUDE FOR REPORT DISTRIBUTION >
    ]

    email_list = str(_select_email(company))
    cc_return = str(_cc_email(company) + cc_list)
    # cc_list = cc_list + cc_return
    outlook = win32.Dispatch('Outlook.Application')
    # outlook.visible = 1
    mail = outlook.CreateItem(0)
    mail.To = email_list
    mail.CC = cc_return
    mail.Subject = 'Sent through Python'
    mail.body = 'This email alert is auto generated. Please do not respond.'
    mail.Attachments.Add(file)
    print(mail.To, mail.CC)
    # mail.send


def ticket_data(tixlist):
    print("Processing....")
    for t in tixlist:
        ticketColList.append(t[0])
        if t[2] is not None:
            typeColList.append(t[2])
        if t[3] is not None:
            priorityColList.append(t[3])
        if t[4] is not None:
            contactColList.append(t[4])

        ticketColSet[t[0]] = tixlist.index(t)
        if t[3] is not None:
            typeColSet[t[2]] = tixlist.index(t)

        if t[3] is not None:
            priorityColSet[t[3]] = tixlist.index(t)

        contactColSet[t[4]] = tixlist.index(t)

    final_priority_set = sorted(priorityColSet)
    # print("tixlist.__len__ = " + str(tixlist.__len__()))
    # print("tixlist[0].__len__ = " + str(tixlist[0].__len__()))

    # tws.max_row = tixlist.__len__()
    # tws.max_column = tixlist[0].__len__()
    row_number = 1
    column_number = 1

    # Company Name
    tws.cell(row=row_number, column=column_number).value = "Company"
    format_heading(row_number, column_number)
    row_number += 1
    # print("ws.cell(2, 1)" + str(ws.cell(2, 1).value))
    tws.cell(row=row_number, column=column_number).value = ws.cell(4, 1).value
    format_title(row_number, column_number)
    row_number += 2

    # Priority Ticket Breakdown
    tws.cell(row=row_number, column=column_number).value = "Total Tickets By Priority"
    format_title(row_number, column_number)
    row_number += 1
    tws.cell(row=row_number, column=column_number).value = "Priority"
    format_heading(row_number, column_number)
    column_number += 1
    tws.cell(row=row_number, column=column_number).value = "Number of Tickets"
    format_heading(row_number, column_number)
    row_number += 1
    column_number = 1

    rows_added = 0
    for fp in final_priority_set:
        # print(tws.cell(row=row_number, column=column_number).value)
        if tws.cell(row=row_number, column=column_number).value is not None:
            row_number += 1
            if str(fp).find('1') != -1:
                tws.cell(row=row_number, column=column_number).value = fp
                format_sub(row_number, column_number)
                column_number += 1
                tws.cell(row=row_number, column=column_number).value = priorityColList.count(fp)
                format_sub(row_number, column_number)
                column_number = 1
                rows_added += 1
            elif str(fp).find('2') != -1:
                tws.cell(row=row_number, column=column_number).value = str(fp)
                format_sub(row_number, column_number)
                column_number += 1
                tws.cell(row=row_number, column=column_number).value = priorityColList.count(fp)
                format_sub(row_number, column_number)
                column_number = 1
                rows_added += 1
            elif str(fp).find('3') != -1:
                tws.cell(row=row_number, column=column_number).value = str(fp)
                format_sub(row_number, column_number)
                column_number += 1
                tws.cell(row=row_number, column=column_number).value = priorityColList.count(fp)
                format_sub(row_number, column_number)
                column_number = 1
                rows_added += 1
            elif str(fp).find('4') != -1:
                tws.cell(row=row_number, column=column_number).value = str(fp)
                format_sub(row_number, column_number)
                column_number += 1
                tws.cell(row=row_number, column=column_number).value = priorityColList.count(fp)
                format_sub(row_number, column_number)
                column_number = 1
                rows_added += 1
        else:
            if str(fp).find('1') != -1:
                tws.cell(row=row_number, column=column_number).value = str(fp)
                format_sub(row_number, column_number)
                column_number += 1
                tws.cell(row=row_number, column=column_number).value = priorityColList.count(fp)
                format_sub(row_number, column_number)
                column_number = 1
                rows_added += 1
            elif str(fp).find('2') != -1:
                tws.cell(row=row_number, column=column_number).value = str(fp)
                format_sub(row_number, column_number)
                column_number += 1
                tws.cell(row=row_number, column=column_number).value = priorityColList.count(fp)
                format_sub(row_number, column_number)
                column_number = 1
                rows_added += 1
            elif str(fp).find('3') != -1:
                tws.cell(row=row_number, column=column_number).value = str(fp)
                format_sub(row_number, column_number)
                column_number += 1
                tws.cell(row=row_number, column=column_number).value = priorityColList.count(fp)
                format_sub(row_number, column_number)
                column_number = 1
                rows_added += 1
            elif str(fp).find('4') != -1:
                tws.cell(row=row_number, column=column_number).value = str(fp)
                format_sub(row_number, column_number)
                column_number += 1
                tws.cell(row=row_number, column=column_number).value = priorityColList.count(fp)
                format_sub(row_number, column_number)
                column_number = 1
                rows_added += 1
    row_number += 1
    column_number += 1

    tws.cell(row=row_number, column=column_number).value = "=SUM(" + str(
        tws.cell((row_number - rows_added), column_number).coordinate) + ":" + str(
        tws.cell((row_number - 1), column_number).coordinate) + ")"
    format_sub_bold(row_number, column_number)
    row_number += 2
    column_number -= 1

    # Type Ticket Breakdown
    rows_added = 0
    tws.cell(row=row_number, column=column_number).value = "Total Tickets By Type"
    format_title(row_number, column_number)
    row_number += 1
    tws.cell(row=row_number, column=column_number).value = "Service Sub Type"
    format_heading(row_number, column_number)
    column_number += 1
    tws.cell(row=row_number, column=column_number).value = "Count"
    format_heading(row_number, column_number)
    row_number += 1
    column_number = 1

    for tc in typeColSet:
        if tc != "Issue Type":
            # print("Type: " + str(tc) + ", Count: " + str(typeColList.count(tc)))
            tws.cell(row=row_number, column=column_number).value = tc
            format_sub(row_number, column_number)
            column_number += 1
            tws.cell(row=row_number, column=column_number).value = typeColList.count(tc)
            format_sub(row_number, column_number)
            row_number += 1
            rows_added += 1
            tws.cell(row=row_number, column=column_number).value = "=SUM(" + str(
                tws.cell((row_number - rows_added), column_number).coordinate) + ":" + str(
                tws.cell((row_number - 1), column_number).coordinate) + ")"
            format_sub_bold(row_number, column_number)
            column_number -= 1

    row_number += 2
    column_number = 1
    # Contact Ticket Breakdown
    rows_added = 0
    tws.cell(row=row_number, column=column_number).value = "Total Tickets By Contact"
    format_title(row_number, column_number)
    row_number += 1
    tws.cell(row=row_number, column=column_number).value = "Contact"
    format_heading(row_number, column_number)
    column_number += 1
    tws.cell(row=row_number, column=column_number).value = "Number of Contacts"
    format_heading(row_number, column_number)
    row_number += 1
    column_number -= 1

    for cc in contactColSet:
        # print("Type: " + str(cc) + ", Count: " + str(contactColList.count(cc)))
        if cc != "Contact":
            tws.cell(row=row_number, column=column_number).value = cc
            format_sub(row_number, column_number)
            column_number += 1
            tws.cell(row=row_number, column=column_number).value = contactColList.count(cc)
            format_sub(row_number, column_number)
            row_number += 1
            rows_added += 1
            tws.cell(row=row_number, column=column_number).value = "=SUM(" + str(
                tws.cell((row_number - rows_added), column_number).coordinate) + ":" + str(
                tws.cell((row_number - 1), column_number).coordinate) + ")"
            format_sub_bold(row_number, column_number)
            column_number -= 1

    row_number += 2
    paste_cells(row_number, (tixlist.__len__()+row_number-1), 1, (tixlist[0].__len__()+1), tws, tixlist)


# VARIABLE SETUP
file_path = r'  <  YOUR FILE PATH  >  '
save_path = r'  <  YOUR FILE PATH  >  '
# This is the path to a text file that contains all company names that should get reports.
company_path = r'  <  YOUR FILE PATH  >  '
company_names = get_company_names(company_path)
report_files = file_array(file_path)

# FORMATTING
font_heading = Font(name="Tahoma", size=9, color='FFFFFF', bold=True)
font_title = Font(name="Calibri", size=14, color="000000")
font_sub = Font(name="Tahoma", size=8)
font_sub_bold = Font(name="Tahoma", size=8, bold=True)
border_all = Side(style="thick", color="000000")
border = Border(left=border_all, right=border_all, top=border_all, bottom=border_all)
fill_heading = PatternFill(fill_type="solid", fgColor="808080")
fill_ticket = PatternFill(fill_type="solid", fgColor="C0C0C0")
alignment_center = Alignment(horizontal="center", wrap_text=True)

# MAIN #
for cn in company_names:
    cn_split = cn.splitlines()
    company_name = cn_split[0]
    for f in report_files:
        if company_name in f:
            ticketColList = []
            typeColList = []
            priorityColList = []
            contactColList = []

            ticketColSet = {}
            typeColSet = {}
            priorityColSet = {}
            contactColSet = {}

            open_path = join(file_path, f)
            ws = open_worksheet(open_path)
            print("ws = " + str(ws))
            twb = Workbook()
            tws = create_worksheet(twb)
            format_cell_width(tws)
            wsRows = ws.max_row
            wsCols = ws.max_column
            newstart = find_text(1, wsRows, 1, wsCols, ws, "Closed")
            ticket_list = copy_cells(newstart, wsRows, 1, wsCols, ws)
            ticket_data(ticket_list)
            save_sheet(twb, save_path, company_name)
            print("the file should be saved with data")


            # Checking if outlook is already opened. If not, open Outlook.exe and send email
            for item in psutil.pids():
                p = psutil.Process(item)
                if p.name() == "OUTLOOK.EXE":
                    flag = 1
                    break
                else:
                    flag = 0

            print(open_path)
            if flag == 1:
                send_notification(open_path, company_name)
            else:
                open_outlook()
                send_notification(open_path, company_name)

            break
        continue
